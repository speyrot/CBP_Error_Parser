import os
from werkzeug.utils import secure_filename
import pandas as pd
from datetime import datetime
from config import UPLOAD_FOLDER, ALLOWED_EXTENSIONS
import logging
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

class FileHandler:
    @staticmethod
    def allowed_file(filename: str, file_type: str) -> bool:
        """
        Check if the file extension is allowed
        
        Args:
            filename (str): Name of the file
            file_type (str): Type of file ('pdf' or 'excel')
            
        Returns:
            bool: True if file extension is allowed, False otherwise
        """
        return '.' in filename and \
            filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS[file_type]

    @staticmethod
    def save_uploaded_file(file, file_type: str) -> str:
        """
        Save an uploaded file to the upload directory
        
        Args:
            file: File object from request
            file_type (str): Type of file ('pdf' or 'excel')
            
        Returns:
            str: Path to the saved file
            
        Raises:
            ValueError: If file type is not allowed
        """
        if file and FileHandler.allowed_file(file.filename, file_type):
            filename = secure_filename(file.filename)
            # Add timestamp to filename to avoid conflicts
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            new_filename = f"{timestamp}_{filename}"
            file_path = os.path.join(UPLOAD_FOLDER, new_filename)
            file.save(file_path)
            return file_path
        raise ValueError(f"Invalid file type. Allowed types for {file_type}: {ALLOWED_EXTENSIONS[file_type]}")

    @staticmethod
    def generate_output_file(df: pd.DataFrame) -> str:
        """
        Generate Excel file from processed data with auto-sized columns
        
        Args:
            df (pd.DataFrame): DataFrame to save
            
        Returns:
            str: Path to the generated Excel file
        """
        logger.debug("Generating output Excel file")
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f"processed_data_{timestamp}.xlsx"
        output_path = os.path.join(UPLOAD_FOLDER, output_filename)
        
        try:
            # Ensure the upload directory exists
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            # Create Excel writer with date format
            writer = pd.ExcelWriter(
                output_path,
                engine='openpyxl',
                date_format='YYYY-MM-DD'
            )
            
            # Save the DataFrame to Excel
            df.to_excel(writer, index=False, sheet_name='Processed Data')
            
            # Access the workbook and active worksheet
            workbook = writer.book
            worksheet = writer.sheets['Processed Data']
            
            # Auto-size columns
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                # Find the maximum length in the column
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Add a little extra width for padding
                adjusted_width = (max_length + 2)
                
                # Set column width
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save the workbook
            writer.close()
            
            logger.info(f"Successfully generated output file: {output_filename}")
            return output_path
            
        except Exception as e:
            logger.exception("Error generating output file")
            raise ValueError(f"Failed to generate output file: {str(e)}")

    @staticmethod
    def cleanup_old_files():
        """Remove files older than 1 hour from the upload directory"""
        # TODO: Implement cleanup logic for old files
        pass
